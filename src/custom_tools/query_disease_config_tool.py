"""
查询疾病详细配置工具
根据疾病ID或疾病名称查询Excel中对应的疾病配置信息
"""

from typing import Any, Dict, Type, Optional
from pydantic import BaseModel, Field
from crewai.tools import BaseTool
import openpyxl
import json
from pathlib import Path


class QueryDiseaseConfigSchema(BaseModel):
    """查询疾病配置工具的输入参数模型"""
    disease_identifiers: str = Field(..., description="疾病名称列表（用逗号分隔）。例如：'肝癌,淋巴瘤' 或 '肝癌'")
    patient_input: Optional[str] = Field(default="", description="可选的患者原始输入文本，用于智能匹配最合适的配置")


class QueryDiseaseConfigTool(BaseTool):
    """
    查询疾病详细配置工具

    功能：
    1. 根据疾病ID或疾病名称查询Excel中的配置信息
    2. 返回该疾病的完整配置，包括检验指标、关键决策指标等

    用途：
    - Agent确定用户的疾病后，调用此工具获取详细配置信息
    - 返回该疾病的所有配置字段
    """

    name: str = "查询疾病配置工具"
    description: str = """
    根据疾病名称查询详细的疾病看板配置信息。支持一次查询多个疾病。

    输入参数：
    - disease_identifiers: 疾病名称列表（用逗号分隔）
      例如：'肝癌,淋巴瘤' 或 '肝癌'
    - patient_input: (可选) 患者原始输入文本，用于更精确的匹配

    返回多个疾病的看板配置列表，每个疾病包括：
    - disease_name: 疾病名称
    - dashboard_description: 看板说明（包含数据提取要求、关键指标、时间轴要求等详细说明）
    """
    args_schema: Type[BaseModel] = QueryDiseaseConfigSchema

    # 缓存数据
    _config_cache: Dict[str, dict] = {}

    def _load_disease_config(self, disease_identifier: str) -> Optional[dict]:
        """
        从Excel加载疾病配置

        参数:
            disease_identifier: 疾病名称

        返回:
            疾病配置字典，如果未找到返回None
        """
        # 检查缓存
        if disease_identifier in self._config_cache:
            return self._config_cache[disease_identifier]

        try:
            # 配置文件路径 - 优先使用新配置文件
            new_config_file = Path(__file__).parent.parent.parent / "app" / "config" / "disease_config_new.xlsx"
            old_config_file = Path(__file__).parent.parent.parent / "app" / "config" / "disease_config.xlsx"

            config_file = new_config_file if new_config_file.exists() else old_config_file

            if not config_file.exists():
                raise Exception(f"配置文件不存在: {config_file}")

            # 读取Excel文件
            wb = openpyxl.load_workbook(config_file)
            ws = wb.active

            # 查找匹配的疾病配置
            for row in range(2, ws.max_row + 1):
                # 新格式：只有disease_name和dashboard_description两列
                if ws.max_column == 2:
                    disease_name = ws.cell(row, 1).value

                    # 支持通过名称匹配
                    if disease_name == disease_identifier:
                        config = {
                            "disease_name": str(disease_name).strip() if disease_name else "",
                            "dashboard_description": str(ws.cell(row, 2).value).strip() if ws.cell(row, 2).value else ""
                        }

                        # 缓存结果
                        self._config_cache[disease_identifier] = config
                        return config
                else:
                    # 旧格式：兼容处理
                    disease_id = ws.cell(row, 1).value
                    disease_name = ws.cell(row, 2).value

                    # 支持通过ID或名称匹配
                    if disease_id == disease_identifier or disease_name == disease_identifier:
                        # 找到匹配的疾病
                        config = {
                            "disease_id": str(disease_id).strip() if disease_id else "",
                            "disease_name": str(disease_name).strip() if disease_name else "",
                            "lab_test_indicators": self._safe_json_load(ws.cell(row, 3).value),
                            "key_decision_indicators": self._safe_json_load(ws.cell(row, 4).value),
                            "highlight_indicators": self._safe_json_load(ws.cell(row, 5).value),
                            "trend_chart_indicators": self._safe_json_load(ws.cell(row, 6).value),
                            "mdt_report_indicators": self._safe_json_load(ws.cell(row, 7).value)
                        }

                        # 缓存结果
                        self._config_cache[disease_identifier] = config
                        return config

            return None

        except Exception as e:
            raise Exception(f"读取疾病配置失败: {str(e)}")

    def _safe_json_load(self, value) -> Any:
        """安全地加载JSON字符串"""
        if not value:
            return {}
        try:
            return json.loads(value)
        except:
            return {}

    def _run(self, disease_identifiers: str, patient_input: str = "", **kwargs: Any) -> Dict:
        """
        执行查询疾病配置工具

        参数:
            disease_identifiers: 疾病名称列表（用逗号分隔）
            patient_input: 患者输入文本（可选）

        返回:
            查询结果字典，包含多个疾病配置的列表
        """
        try:
            if not disease_identifiers:
                return {
                    "status": "error",
                    "message": "疾病名称不能为空",
                    "configs": []
                }

            # 分割疾病标识符列表
            identifier_list = [id.strip() for id in disease_identifiers.split(',') if id.strip()]

            if not identifier_list:
                return {
                    "status": "error",
                    "message": "没有有效的疾病名称",
                    "configs": []
                }

            # 查询每个疾病的配置
            configs = []
            not_found = []

            for identifier in identifier_list:
                config = self._load_disease_config(identifier)
                if config:
                    configs.append(config)
                else:
                    not_found.append(identifier)

            # 如果没有找到任何配置，尝试加载通用配置
            if not configs:
                general_config = self._load_disease_config("通用")
                if general_config:
                    return {
                        "status": "fallback_to_general",
                        "message": f"未找到指定疾病的配置: {', '.join(not_found)}，返回通用配置",
                        "not_found_diseases": not_found,
                        "configs": [general_config]
                    }
                else:
                    return {
                        "status": "not_found",
                        "message": f"未找到任何疾病配置: {', '.join(not_found)}",
                        "not_found_diseases": not_found,
                        "configs": []
                    }

            # 返回成功结果
            result = {
                "status": "success",
                "count": len(configs),
                "configs": configs
            }

            # 如果有部分疾病未找到，添加警告信息
            if not_found:
                result["warning"] = f"部分疾病未找到配置: {', '.join(not_found)}"
                result["not_found_diseases"] = not_found

            return result

        except Exception as e:
            return {
                "status": "error",
                "message": f"查询失败: {str(e)}",
                "configs": []
            }


# 创建工具实例
query_disease_config_tool = QueryDiseaseConfigTool()
