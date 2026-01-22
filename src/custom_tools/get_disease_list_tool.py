"""
获取疾病列表工具
从Excel配置文件中读取并返回所有疾病列表
"""

from typing import Any, Type, List, Optional
from pydantic import BaseModel, Field
from crewai.tools import BaseTool
import openpyxl
from pathlib import Path


class GetDiseaseListSchema(BaseModel):
    """获取疾病列表工具的输入参数模型"""
    dummy: Optional[str] = Field(default="", description="占位参数，本工具不需要输入参数")


class GetDiseaseListTool(BaseTool):
    """
    获取疾病列表工具

    功能：
    1. 读取疾病配置Excel文件
    2. 提取"疾病名称"列的所有疾病
    3. 返回疾病列表

    用途：
    - Agent首先调用此工具获取所有可用的疾病
    - 然后根据用户输入判断最接近的疾病
    """

    name: str = "获取疾病列表工具"
    description: str = """
    获取所有可用的疾病列表。
    返回Excel配置文件中所有疾病的名称。
    此工具不需要任何输入参数，直接调用即可。
    """
    args_schema: Type[BaseModel] = GetDiseaseListSchema

    # 缓存数据
    _disease_list_cache: Optional[List[dict]] = None

    def _load_disease_list(self) -> List[dict]:
        """
        从Excel加载疾病列表并缓存

        返回:
            疾病列表，每项包含 disease_name
        """
        if self._disease_list_cache is not None:
            return self._disease_list_cache

        try:
            # 配置文件路径
            config_file = Path(__file__).parent.parent.parent / "app" / "config" / "disease_config.xlsx"

            if not config_file.exists():
                raise Exception(f"配置文件不存在: {config_file}")

            # 读取Excel文件
            wb = openpyxl.load_workbook(config_file)
            ws = wb.active

            diseases = []

            # 从第2行开始读取（第1行是表头）
            for row in range(2, ws.max_row + 1):
                # 新格式：只有disease_name和dashboard_description两列
                if ws.max_column == 2:
                    disease_name = ws.cell(row, 1).value
                    if disease_name:
                        diseases.append({
                            "disease_name": str(disease_name).strip()
                        })
                else:
                    # 旧格式：兼容处理
                    disease_id = ws.cell(row, 1).value
                    disease_name = ws.cell(row, 2).value

                    if disease_id and disease_name:
                        diseases.append({
                            "disease_id": str(disease_id).strip(),
                            "disease_name": str(disease_name).strip()
                        })

            # 缓存结果
            self._disease_list_cache = diseases
            return diseases

        except Exception as e:
            raise Exception(f"读取疾病配置文件失败: {str(e)}")

    def _run(self, **kwargs: Any) -> dict:
        """
        执行获取疾病列表工具

        返回:
            包含疾病列表的字典
        """
        try:
            # 加载疾病列表
            diseases = self._load_disease_list()

            return {
                "status": "success",
                "count": len(diseases),
                "diseases": diseases
            }

        except Exception as e:
            return {
                "status": "error",
                "message": f"获取疾病列表失败: {str(e)}",
                "diseases": []
            }


# 创建工具实例
get_disease_list_tool = GetDiseaseListTool()
